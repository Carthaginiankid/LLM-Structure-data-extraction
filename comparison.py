from typing import List, Dict, Optional
from models import ExtractedQuotation, Recommendation
from llm_extractor import LLMExtractor
import pandas as pd
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

EXCHANGE_RATES = {"USD": 0.92, "EUR": 1.0, "GBP": 1.17, "JPY": 0.0062}

def convert_to_eur(amount: float, from_currency: str) -> float:
    if from_currency == "EUR":
        return amount
    return amount * EXCHANGE_RATES.get(from_currency, 1.0)

class ProcurementScorer:
    WEIGHTS = {"tco": 0.35, "delivery": 0.25, "payment": 0.20, "tooling": 0.10, "moq": 0.10}

    def score_all(self, comparison_data: List[Dict]) -> List[Dict]:
        if not comparison_data:
            return []
        for item in comparison_data:
            item["scores"] = self._calculate_all_scores(item, comparison_data)
            item["total_score"] = self._calculate_weighted_total(item["scores"])
        scored_data = sorted(comparison_data, key=lambda x: x["total_score"], reverse=True)
        for idx, item in enumerate(scored_data, 1):
            item["final_ranking"] = idx
        return scored_data

    def _calculate_all_scores(self, item: Dict, all_items: List[Dict]) -> Dict:
        return {
            "tco_score": self._score_tco(item, all_items),
            "delivery_score": self._score_delivery(item, all_items),
            "payment_score": self._score_payment(item, all_items),
            "tooling_score": self._score_tooling(item, all_items),
            "moq_score": self._score_moq(item, all_items),
            "missing_data_penalty": self._calculate_missing_data_penalty(item)
        }

    def _score_tco(self, item: Dict, all_items: List[Dict]) -> float:
        tcos = [d["total_cost_eur"] for d in all_items]
        min_tco, max_tco = min(tcos), max(tcos)
        if max_tco == min_tco:
            return 100.0
        score = 100 * (1 - (item["total_cost_eur"] - min_tco) / (max_tco - min_tco))
        return round(max(0, score), 2)

    def _score_delivery(self, item: Dict, all_items: List[Dict]) -> float:
        lead_times = [d.get("lead_time_weeks") for d in all_items if d.get("lead_time_weeks") is not None]
        if not lead_times or item.get("lead_time_weeks") is None:
            return 20.0
        min_weeks, max_weeks = min(lead_times), max(lead_times)
        if max_weeks == min_weeks:
            return 100.0
        score = 100 * (1 - (item["lead_time_weeks"] - min_weeks) / (max_weeks - min_weeks))
        return round(max(0, score), 2)

    def _score_payment(self, item: Dict, all_items: List[Dict]) -> float:
        payment_days_list = [d.get("payment_days") for d in all_items if d.get("payment_days") is not None]
        if not payment_days_list or item.get("payment_days") is None:
            return 20.0
        min_days, max_days = min(payment_days_list), max(payment_days_list)
        if max_days == min_days:
            return 100.0
        score = 100 * (1 - (item["payment_days"] - min_days) / (max_days - min_days))
        return round(max(0, score), 2)

    def _score_tooling(self, item: Dict, all_items: List[Dict]) -> float:
        tooling_costs = [d["tooling_cost_eur"] for d in all_items]
        min_tooling, max_tooling = min(tooling_costs), max(tooling_costs)
        if max_tooling == min_tooling:
            return 100.0
        score = 100 * (1 - (item["tooling_cost_eur"] - min_tooling) / (max_tooling - min_tooling))
        return round(max(0, score), 2)

    def _score_moq(self, item: Dict, all_items: List[Dict]) -> float:
        moqs = [d.get("moq") for d in all_items if isinstance(d.get("moq"), (int, float)) and d.get("moq") > 0]
        if not moqs or not isinstance(item.get("moq"), (int, float)) or item.get("moq") <= 0:
            return 20.0
        min_moq, max_moq = min(moqs), max(moqs)
        if max_moq == min_moq:
            return 100.0
        score = 100 * (1 - (item["moq"] - min_moq) / (max_moq - min_moq))
        return round(max(0, score), 2)

    def _calculate_missing_data_penalty(self, item: Dict) -> float:
        missing_fields = [
            item.get("lead_time") in [None, "N/A", ""],
            item.get("payment_terms") in [None, "N/A", ""],
            item.get("delivery_terms") in [None, "N/A", ""],
            item.get("quotation_date") in [None, "N/A", ""],
            item.get("moq") in [None, "N/A", 0],
            item.get("tooling_cost_eur", 0) == 0 and item.get("tooling_cost_original", 0) == 0
        ]
        return min(100, sum(missing_fields) * 10)

    def _calculate_weighted_total(self, scores: Dict) -> float:
        base_total = (
            scores["tco_score"] * self.WEIGHTS["tco"] +
            scores["delivery_score"] * self.WEIGHTS["delivery"] +
            scores["payment_score"] * self.WEIGHTS["payment"] +
            scores["tooling_score"] * self.WEIGHTS["tooling"] +
            scores["moq_score"] * self.WEIGHTS["moq"]
        )
        penalty_factor = 1 - (scores.get("missing_data_penalty", 0) / 100)
        return round(max(0, base_total * penalty_factor), 2)

class LLMRecommendationEngine:
    def __init__(self, provider: str = "groq", model: str = "llama-3.3-70b-versatile"):
        self.extractor = LLMExtractor(provider=provider, model=model)

    def generate_recommendation(self, comparison_data: List[Dict]) -> Dict:
        if not comparison_data:
            return {
                "recommended_supplier": None,
                "total_score": 0,
                "reasoning": "No suppliers available.",
                "key_advantages": [],
                "considerations": []
            }
        context = self._prepare_context(comparison_data)
        recommendation = self.extractor.client.chat.completions.create(
            model=self.extractor.model,
            response_model=Recommendation,
            messages=[
                {
                    "role": "system",
                    "content": "You are a senior procurement analyst with 15+ years of experience. Provide data-driven supplier recommendations with specific quantitative comparisons."
                },
                {
                    "role": "user",
                    "content": f"""Analyze the supplier comparison data and recommend the supplier with the highest total score.

{context}

Provide:
1. Detailed reasoning (400-600 words) comparing the recommended supplier against each competitor with specific metrics (cost differences in EUR, delivery times, payment terms, tooling costs, score differences)
2. 4-6 key advantages with specific metrics
3. 3-5 considerations or trade-offs
4. Note any missing data impact if applicable

Be specific with numbers, percentages, EUR amounts, and timeframes. Use professional procurement terminology."""
                }
            ],
            temperature=0.4
        )
        return {
            "recommended_supplier": recommendation.recommended_supplier or comparison_data[0]["supplier"],
            "total_score": recommendation.total_score or comparison_data[0].get("total_score", 0),
            "reasoning": recommendation.reasoning,
            "key_advantages": recommendation.key_advantages or [],
            "considerations": recommendation.considerations or [],
            "missing_data_impact": recommendation.missing_data_impact
        }

    def _prepare_context(self, comparison_data: List[Dict]) -> str:
        lines = ["SUPPLIER COMPARISON DATA:\n", "=" * 80]
        for idx, s in enumerate(comparison_data, 1):
            scores = s.get("scores", {})
            lines.append(f"\nSupplier {idx}: {s['supplier']} (Rank: {s.get('ranking', idx)}, Score: {s.get('total_score', 0):.1f}/100)")
            lines.append(f"  TCO: €{s['total_cost_eur']:,.2f} | Tooling: €{s['tooling_cost_eur']:,.2f} | Avg Unit: €{s['unit_cost_avg_eur']:.2f}")
            lines.append(f"  Scores: TCO {scores.get('tco_score', 0):.1f} | Delivery {scores.get('delivery_score', 0):.1f} | Payment {scores.get('payment_score', 0):.1f} | Tooling {scores.get('tooling_score', 0):.1f} | MOQ {scores.get('moq_score', 0):.1f}")
            lines.append(f"  Lead Time: {s.get('lead_time', 'N/A')} ({s.get('lead_time_weeks', 'N/A')} weeks) | Payment: {s.get('payment_terms', 'N/A')} ({s.get('payment_days', 'N/A')} days) | MOQ: {s.get('moq', 'N/A')}")
        return "\n".join(lines)

class QuotationComparator:
    def compare(self, quotations: List[ExtractedQuotation]) -> dict:
        if not quotations:
            return {}
        comparison_data = [self._build_comparison_item(q) for q in quotations]
        scored_data = ProcurementScorer().score_all(comparison_data)
        comparison_data.sort(key=lambda x: x["total_cost_eur"])
        for idx, item in enumerate(comparison_data, 1):
            item["cost_ranking"] = idx
        for item in scored_data:
            item["ranking"] = item.get("final_ranking", item.get("cost_ranking", 999))
        recommendation = LLMRecommendationEngine().generate_recommendation(scored_data) if scored_data else None
        return {
            "comparison_table": scored_data,
            "summary": self._generate_summary(scored_data),
            "recommendation": recommendation,
            "generated_at": datetime.now().isoformat()
        }

    def _build_comparison_item(self, quote: ExtractedQuotation) -> Dict:
        tco_original = self._calculate_tco(quote)
        tooling_base = self._calculate_tooling_cost(quote)
        price_breakdown_original = self._get_price_breakdown_original(quote)
        return {
            "supplier": quote.supplier_name,
            "original_currency": quote.currency.value,
            "total_cost_eur": convert_to_eur(tco_original, quote.currency.value),
            "total_cost_original": tco_original,
            "tooling_cost_eur": convert_to_eur(tooling_base, quote.currency.value),
            "tooling_cost_original": tooling_base,
            "tooling_cost_type": getattr(quote, 'tooling_cost_type', None),
            "unit_cost_avg_eur": self._calculate_avg_unit_cost_eur(quote),
            "unit_cost_avg_original": self._calculate_avg_unit_cost(quote),
            "delivery_terms": quote.delivery_terms or "N/A",
            "incoterms": self._extract_incoterms(quote.delivery_terms),
            "lead_time": quote.lead_time or "N/A",
            "lead_time_weeks": self._parse_lead_time(quote.lead_time),
            "payment_terms": quote.payment_terms or "N/A",
            "payment_days": self._parse_payment_terms(quote.payment_terms),
            "moq": quote.moq or "N/A",
            "quotation_date": quote.quotation_date.strftime("%Y-%m-%d") if quote.quotation_date else "N/A",
            "price_breakdown_eur": self._get_price_breakdown_eur(quote),
            "price_breakdown_original": price_breakdown_original,
            "years_covered": list(quote.annual_prices.keys()) if quote.annual_prices else []
        }

    def _calculate_tooling_cost(self, quote: ExtractedQuotation) -> float:
        tooling_base = quote.tooling_cost or 0
        tooling_cost_type = getattr(quote, 'tooling_cost_type', None)
        if tooling_cost_type and tooling_cost_type.lower() in ['renewal', 'recurring']:
            tooling_base *= len(quote.annual_prices) if quote.annual_prices else 1
        return tooling_base

    def _calculate_tco(self, quote: ExtractedQuotation) -> float:
        total = sum(quote.annual_prices[y] * quote.annual_quantities.get(y, 0) for y in quote.annual_prices.keys())
        if quote.tooling_cost:
            tooling_cost_type = getattr(quote, 'tooling_cost_type', None)
            if tooling_cost_type and tooling_cost_type.lower() in ['renewal', 'recurring']:
                total += quote.tooling_cost * (len(quote.annual_prices) if quote.annual_prices else 1)
            else:
                total += quote.tooling_cost
        return total

    def _calculate_avg_unit_cost(self, quote: ExtractedQuotation) -> float:
        if not quote.annual_prices:
            return 0.0
        return sum(quote.annual_prices.values()) / len(quote.annual_prices)

    def _calculate_avg_unit_cost_eur(self, quote: ExtractedQuotation) -> float:
        if not quote.annual_prices:
            return 0.0
        return sum(convert_to_eur(p, quote.currency.value) for p in quote.annual_prices.values()) / len(quote.annual_prices)

    def _get_price_breakdown_original(self, quote: ExtractedQuotation) -> dict:
        breakdown = {}
        for year in sorted(quote.annual_prices.keys()):
            unit_price = quote.annual_prices[year]
            quantity = quote.annual_quantities.get(year, 0)
            breakdown[year] = {
                "unit_price": unit_price,
                "quantity": quantity,
                "total": unit_price * quantity
            }
        return breakdown

    def _get_price_breakdown_eur(self, quote: ExtractedQuotation) -> dict:
        breakdown = {}
        for year in sorted(quote.annual_prices.keys()):
            unit_price_eur = convert_to_eur(quote.annual_prices[year], quote.currency.value)
            quantity = quote.annual_quantities.get(year, 0)
            breakdown[year] = {
                "unit_price": unit_price_eur,
                "quantity": quantity,
                "total": unit_price_eur * quantity
            }
        return breakdown

    def _parse_lead_time(self, lead_time: Optional[str]) -> Optional[float]:
        if not lead_time or lead_time == "N/A":
            return None
        try:
            numbers = re.findall(r'\d+\.?\d*', lead_time.lower())
            if not numbers:
                return None
            weeks = float(numbers[0])
            lead_time_lower = lead_time.lower()
            if 'day' in lead_time_lower or 'tag' in lead_time_lower:
                weeks /= 7
            elif 'month' in lead_time_lower or 'monat' in lead_time_lower:
                weeks *= 4.33
            elif 'year' in lead_time_lower or 'jahr' in lead_time_lower:
                weeks *= 52
            return round(weeks, 1)
        except (ValueError, TypeError):
            return None

    def _parse_payment_terms(self, payment_terms: Optional[str]) -> Optional[int]:
        if not payment_terms or payment_terms == "N/A":
            return None
        try:
            match = re.search(r'(\d+)\s*(?:days?|tag)', payment_terms.lower()) or re.search(r'net\s*(\d+)', payment_terms.lower())
            return int(match.group(1)) if match else None
        except (ValueError, AttributeError):
            return None

    def _extract_incoterms(self, delivery_terms: Optional[str]) -> str:
        if not delivery_terms or delivery_terms == "N/A":
            return "N/A"
        incoterms = ["EXW", "FCA", "FAS", "FOB", "CFR", "CIF", "CPT", "CIP", "DAP", "DPU", "DDP"]
        return next((term for term in incoterms if term in delivery_terms.upper()), "N/A")

    def _generate_summary(self, comparison_data: List[dict]) -> dict:
        if not comparison_data:
            return {}
        costs = [d["total_cost_eur"] for d in comparison_data]
        return {
            "total_suppliers": len(comparison_data),
            "lowest_cost": min(costs),
            "highest_cost": max(costs),
            "cost_range": max(costs) - min(costs),
            "best_supplier": comparison_data[0].get("supplier") if comparison_data else None,
            "best_score": comparison_data[0].get("total_score", 0) if comparison_data else 0
        }

    def to_dataframe(self, comparison: dict) -> pd.DataFrame:
        if not comparison or "comparison_table" not in comparison:
            return pd.DataFrame()
        rows = []
        for item in comparison["comparison_table"]:
            scores = item.get("scores", {})
            price_summary = self._format_price_summary(item["price_breakdown_eur"])
            rec_note = "RECOMMENDED" if item["ranking"] == 1 and comparison.get("recommendation", {}).get("recommended_supplier") == item["supplier"] else ""
            rows.append({
                "Rank": item["ranking"],
                "Supplier Name": item["supplier"],
                "Recommendation": rec_note,
                "Total Score": f"{item.get('total_score', 0):.1f}",
                "Missing Data Penalty": f"-{scores.get('missing_data_penalty', 0):.1f}" if scores.get('missing_data_penalty', 0) > 0 else "0",
                "TCO Score (35%)": f"{scores.get('tco_score', 0):.1f}",
                "Delivery Score (25%)": f"{scores.get('delivery_score', 0):.1f}",
                "Payment Score (20%)": f"{scores.get('payment_score', 0):.1f}",
                "Tooling Score (10%)": f"{scores.get('tooling_score', 0):.1f}",
                "MOQ Score (10%)": f"{scores.get('moq_score', 0):.1f}",
                "Total Cost (TCO) - EUR": f"€{item['total_cost_eur']:,.2f}",
                "Tooling Cost - EUR": f"€{item['tooling_cost_eur']:,.2f}",
                "Average Unit Price - EUR": f"€{item['unit_cost_avg_eur']:.2f}",
                "Price Summary (EUR)": price_summary,
                "Original Currency": item["original_currency"],
                "Total Cost (Original)": f"{item['total_cost_original']:,.2f} {item['original_currency']}",
                "Incoterms": item["incoterms"],
                "Delivery Terms": item["delivery_terms"],
                "Lead Time": item["lead_time"],
                "Lead Time (Weeks)": item.get("lead_time_weeks", "N/A"),
                "Payment Terms": item["payment_terms"],
                "Payment Days": item["payment_days"] if item.get("payment_days") else "N/A",
                "MOQ": item["moq"],
                "Quotation Date": item["quotation_date"],
                "Years Covered": ", ".join(map(str, sorted(item["years_covered"])))
            })
        return pd.DataFrame(rows)

    def _format_price_summary(self, breakdown: dict) -> str:
        if not breakdown:
            return "N/A"
        years = sorted(breakdown.keys())
        if not years:
            return "N/A"
        if len(years) == 1:
            return f"{years[0]}: €{breakdown[years[0]]['unit_price']:.2f}"
        first_price, last_price = breakdown[years[0]]["unit_price"], breakdown[years[-1]]["unit_price"]
        trend = "↓" if last_price < first_price else "↑" if last_price > first_price else "→"
        reduction = abs((first_price - last_price) / first_price * 100) if first_price > 0 else 0
        return f"{years[0]}: €{first_price:.2f} → {years[-1]}: €{last_price:.2f} ({trend} {reduction:.1f}%)"

    def export_to_excel(self, comparison: dict, filename: str = "comparison.xlsx"):
        if not comparison or "comparison_table" not in comparison:
            print("No data to export")
            return
        main_df = self.to_dataframe(comparison)
        rec = comparison.get("recommendation", {})
        summary = comparison.get("summary", {})
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            main_df.to_excel(writer, sheet_name="Supplier Comparison", index=False)
            if rec:
                pd.DataFrame({
                    "Section": ["Recommended Supplier", "Total Score", "Reasoning", "Key Advantages", "Considerations"],
                    "Details": [
                        rec.get("recommended_supplier", "N/A"),
                        f"{rec.get('total_score', 0):.1f}/100",
                        rec.get("reasoning", "N/A"),
                        "; ".join(rec.get("key_advantages", [])) or "N/A",
                        "; ".join(rec.get("considerations", [])) or "N/A"
                    ]
                }).to_excel(writer, sheet_name="Recommendation", index=False)
            pd.DataFrame({
                "Metric": [
                    "Total Suppliers", "Recommended Supplier (Highest Score)", "Total Score",
                    "Lowest Total Cost (EUR)", "Highest Total Cost (EUR)", "Cost Range (EUR)",
                    "Recommendation Reasoning", "Key Advantages", "Considerations",
                    "Scoring Methodology", "Generated At"
                ],
                "Value": [
                    summary.get("total_suppliers", 0),
                    summary.get("best_supplier", "N/A"),
                    f"{rec.get('total_score', 0):.1f}/100",
                    f"€{summary.get('lowest_cost', 0):,.2f}",
                    f"€{summary.get('highest_cost', 0):,.2f}",
                    f"€{summary.get('cost_range', 0):,.2f}",
                    (rec.get("reasoning", "N/A")[:500] + "...") if len(rec.get("reasoning", "")) > 500 else rec.get("reasoning", "N/A"),
                    "; ".join(rec.get("key_advantages", [])) or "N/A",
                    "; ".join(rec.get("considerations", [])) or "N/A",
                    "Weighted scoring: TCO(35%) + Delivery(25%) + Payment(20%) + Tooling(10%) + MOQ(10%)",
                    comparison.get("generated_at", "N/A")
                ]
            }).to_excel(writer, sheet_name="Summary", index=False)
        self._format_excel_file(filename)

    def _format_excel_file(self, filename: str):
        try:
            wb = load_workbook(filename)
            header_style = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for sheet_name in ["Supplier Comparison", "Summary"]:
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                for cell in ws[1]:
                    cell.fill = header_style
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = border
                for column in ws.columns:
                    max_length = max((len(str(cell.value)) for cell in column if cell.value), default=0)
                    adjusted_width = min(max_length + 2, 80 if sheet_name == "Summary" else 50)
                    ws.column_dimensions[column[0].column_letter].width = adjusted_width
                if sheet_name == "Supplier Comparison":
                    ws.freeze_panes = "B2"
            wb.save(filename)
            print(f"Exported to: {filename}")
        except Exception as e:
            print(f"Warning: Could not format Excel file: {e}")
